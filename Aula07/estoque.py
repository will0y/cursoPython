import openpyxl
from openpyxl.styles import Font

def criar_planilha():
    # Cria um novo arquivo Excel
    planilha = openpyxl.Workbook()

    # Seleciona a planilha ativa (por padrão, a primeira planilha é criada)
    planilha_ativa = planilha.active

    # Define os títulos das colunas
    planilha_ativa.append(['Nome do Produto', 'Quantidade', 'Marca', 'Valor'])

    # Aplica negrito aos títulos das colunas
    for cell in planilha_ativa[1]:
        cell.font = Font(bold=True)

    # Salva o arquivo
    planilha.save('controle_estoque.xlsx')

def adicionar_produto(nome, quantidade, marca, valor):
    # Abre o arquivo Excel existente
    planilha = openpyxl.load_workbook('controle_estoque.xlsx')

    # Seleciona a planilha ativa
    wb = planilha.active

    # Adiciona os dados do produto
    wb.append([nome, quantidade, marca, valor])

    # Salva o arquivo
    planilha.save('controle_estoque.xlsx')

def excluir_produto(nome, quantidade, marca):
    planilha = openpyxl.load_workbook('controle_estoque.xlsx')
    wb = planilha.active
    
    for row in wb.iter_rows(min_row=2, values_only=True):
        if row[0] == nome and row[1] == quantidade and row[2] == marca:
            wb.delete_rows(row[0])
            break
    
    planilha.save('controle_estoque.xlsx')

def listar_estoque():
    planilha = openpyxl.load_workbook('controle_estoque.xlsx')
    wb = planilha.active

    print("\nEstoque:")
    for row in wb.iter_rows(min_row=2, values_only=True):
        print("Nome:", row[0])
        print("Quantidade:", row[1])
        print("Marca:", row[2])
        print("Valor:", row[3])
        print("-" * 30)

def main():
    print("Controle de Estoque de Eletrônicos")

    while True:
        print("\nOpções:")
        print("1. Criar planilha de estoque")
        print("2. Adicionar produto")
        print("3. Excluir produto")
        print("4. Listar estoque")
        print("5. Sair")

        opcao = input("Escolha uma opção: ")

        if opcao == '1':
            criar_planilha()
            print("Planilha criada com sucesso!")

        elif opcao == '2':
            nome = input("Nome do Produto: ")
            quantidade = int(input("Quantidade: "))
            marca = input("Marca: ")
            valor = float(input("Valor: "))

            adicionar_produto(nome, quantidade, marca, valor)
            print("Produto adicionado com sucesso!")

        elif opcao == '3':
            nome = input("Digite o nome do produto a ser excluído:")
            quantidade = int(input("Digite a quantidade do produto a ser excluído:"))
            marca = input("Digite a marca do produto a ser excluído:")
            
            excluir_produto(nome, quantidade, marca)
            print("Produto excluído com sucesso!")

        elif opcao == '4':
            listar_estoque()

        elif opcao == '5':
            print("Encerrando o programa.")
            break

        else:
            print("Opção inválida. Por favor, escolha uma opção válida.")


if __name__ == "__main__":
    main()



    #A linha de código if __name__ == "__main__":
    # 
    #  é usada para verificar se o script Python está sendo executado como um programa independente ou se está sendo importado como um módulo em outro script.#
