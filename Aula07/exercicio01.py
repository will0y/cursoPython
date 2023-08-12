#Exercício 1: Criar uma planilha e adicionar dados

import openpyxl

# Criar uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active

# Adicionar dados às células
ws['A1'] = 'Nome'
ws['B1'] = 'Idade'
ws['A2'] = 'Alice'
ws['B2'] = 28
ws['A3'] = 'Bob'
ws['B3'] = 32

# Salvar a planilha
wb.save('exercicio1.xlsx')


