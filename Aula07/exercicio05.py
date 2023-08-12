#Exercício 5: Formatar células

import openpyxl
from openpyxl.styles import Font

# Carregar a planilha existente
wb = openpyxl.load_workbook('exercicio4.xlsx')
ws = wb.active

# Formatar células
bold_font = Font(bold=True)
ws['A1'].font = bold_font
ws['B1'].font = bold_font
ws['C1'].font = bold_font
ws['A4'].font = bold_font


# Salvar a planilha atualizada
wb.save('exercicio5.xlsx')

# Fechar a planilha
wb.close()
