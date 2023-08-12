#Exercicio 2: ler dados nde uma planilha 

import openpyxl

#Carregar a planilha existente 
wb = openpyxl.load_workbook('exercicio01.xlsx')
ws = wb.active

#Ler dados de células 
nome = ws['A2'], value
idade = ws['B2'], value 
print(f'Nome: (nome), Idade: (idade)')

#Fechar a planilha
wb.close()