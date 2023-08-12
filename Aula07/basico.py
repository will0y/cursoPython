import openpyxl
from openpyxl.styles import Font

# Criar uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active

# Adicionar dados às células
ws['A1'] = 'Nome'
ws['B1'] = 'Idade'
ws['A2'] = 'Alice'
ws['B2'] = 28
ws['A3'] = 'Bob'
ws['B3'] = 32

# Formatar células
bold_font = Font(bold=True)
ws['A1'].font = bold_font
ws['B1'].font = bold_font

# Salvar planilha em um arquivo
wb.save('exemplo.xlsx')

# Carregar uma planilha existente
wb = openpyxl.load_workbook('exemplo.xlsx')
ws = wb.active

# Ler dados de células
nome = ws['A2'].value
idade = ws['B2'].value
print(f'Nome: {nome}, Idade: {idade}')

# Atualizar valores de células
ws['B3'] = 33

# Adicionar uma nova coluna
ws['C1'] = 'Cargo'
ws['C2'] = 'Engenheiro'
ws['C3'] = 'Designer'

# Salvar novamente a planilha atualizada
wb.save('exemplo_atualizado.xlsx')

# Fechar a planilha
wb.close()
