#Exercício 4: Adicionar uma nova coluna e calcular médias

import openpyxl

# Carregar a planilha existente
wb = openpyxl.load_workbook('exercicio3.xlsx')
ws = wb.active

# Adicionar uma nova coluna
ws['A4']= 'Média'
ws['C1'] = 'Salário'
ws['C2'] = 5000
ws['C3'] = 6000

# Calcular a média dos salários
media = (ws['C2'].value + ws['C3'].value) / 2
ws['C4'] = media

#Calcular a média da idade
media_idade = (ws['B2'].value + ws['B3'].value) / 2
ws['B4'] = media_idade

# Salvar a planilha atualizada
wb.save('exercicio4.xlsx')

# Fechar a planilha
wb.close()
