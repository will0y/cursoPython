#Exercício 2: Ler dados de uma planilha

import openpyxl

# Carregar a planilha existente
wb = openpyxl.load_workbook('exercicio1.xlsx')
ws = wb.active

# Ler dados de células
nome = ws['A2'].value
idade = ws['B2'].value
print(f'Nome: {nome}, Idade: {idade}')

#wb.save('lendo_planilha.xlsx')

# Fechar a planilha
wb.close()
