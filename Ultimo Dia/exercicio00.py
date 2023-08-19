import openpyxl 

wb = openpyxl.Workbook()
ws = wb.active()

ws['A1'] = 'Brasil'
ws['A2'] = '21100000'
ws['B1'] = 'Estados Unidos'
ws['B2'] = '33100000'
ws['C1'] = 'China'
ws['C2'] = '1412000000' 

wb.save('exercicio00.xlsx') 