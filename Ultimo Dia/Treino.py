import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = 'Nome'
ws['B1'] = 'Idade'
ws['C1'] = 'Profissão'
ws['A2'] = 'Steve'
ws['B2'] = 21
ws['C3'] = 'Empresário'
 
Planilha = wb,ws
for planilha in ws:
    print("lala")

bold_font = Font(bold=True)
ws['A1'].font = bold_font
ws['B1'].font = bold_font


wb.save('exemplo.xlsx')
